import discord
client = discord.Client()
import openpyxl
import random
from captcha. image import ImageCaptcha
import os

@client.event

async def on_ready():
    print("Ready")
    game = discord.Game("실험체를 생성중...")
    await client.change_presence(status=discord.Status.online, activity=game)

@client.event

async def on_message(message):

    HelloArr = ["안녕!" + message.author, "나불렀썽?", "짠 슈리 등장! :laughing:"]

    if message.content.startswith("=도움말"):
        author = message.author
        embed = discord.Embed(color=0x0000ff)
        #embed.add_field(name="공지", value="공지를 올립니다.\n사용법 : -공지 \"할말\"")
        await author.send(embed=embed)

    if message.content.startswith("=안녕"):
        await message.channel.send(HelloArr[int(random.randint(0, 2))])

    if message.content.startswith("=공지"):
        if message.content[4:] != "":
            msg = message.content.split(" ")[1]
            await message.channel.send("@everyone " + msg)
        else:
            await message.channel.send("공지글을 입력해 주세요")

    if message.content.startswith("") and message.author.id != 605338386513002526 and message.channel != "dm":
        file = openpyxl.load_workbook("level.xlsx")
        sheet = file.active
        exp = [10, 25, 40, 70, 120, 200, 1000000000]
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + int(random.randint(1, 7))
                if sheet["B" + str(i)].value >= exp[sheet["C" + str(i)].value - 1]:
                    sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1
                    await message.channel.send("레벨 업! (레벨 : " + str(sheet["C" + str(i)].value) + ")")
                file.save("level.xlsx")
                break

            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 0
                sheet["C" + str(i)].value = 1
                file.save("level.xlsx")
                break

            i += 1
            
            #레벨링 시스템 감시하면서 print로 로그 사용하여 확인하기 + 랜덤변수 확인하기

    if message.content == ("=관리자"):
        file = openpyxl.load_workbook("GM.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.guild.id):
                await message.channel.send("현재 봇 관리자는 " + str(sheet["C" + str(i)].value) + "입니다.")
                file.save("GM.xlsx")
                break
            if sheet["A" + str(i)].value == None:
                await message.channel.send("설정된 관리자가 없습니다. \"=관리자 설정 (유저 아이디) (이름)\" 을 사용하여 관리자를 설정해 주세요")
                if sheet["A" + str(i)].value == "":
                    sheet["A" + str(i)].value = str(message.guild.id)
                file.save("GM.xlsx")
                break

            i += 1

    if message.content.startswith("=관리자 설정"):
        file = openpyxl.load_workbook("GM.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.guild.id):
                if sheet["B" + str(i)].value == str(message.author.id) or sheet["B" + str(i)].value == "":
                    userid = message.content[8:26]
                    name = message.content[27:]
                    if id == "" or name == "":
                        await message.channel.send("유저 아이디 또는 유저 이름을 지정해 주세요")
                        break
                    else:
                        sheet["B" + str(i)].value = str(userid)
                        sheet["C" + str(i)].value = str(name)
                        await message.channel.send("관리자가 설정되었습니다.")
                        file.save("GM.xlsx")
                        break
                else:
                    await message.channel.send("관리자가 아닙니다. 설정을 변경하려면 관리자에게 문의하세요.")
                    break
        i += 1

    if message.channel == "dm":
        await message.author.send("Hi")

    if message.content.startswith('=연산'):
        if (0 == 0):
            Member = message.author
            Resurt = 0
            FirstNum = int(random.randint(0, 100))
            SecondNum = int(random.randint(1, 100))
            CalcArr = ["+", "-", "×", "÷"]
            Calc = CalcArr[int(random.randint(0, 3))]
            if (Calc == "+"):
                Resurt = FirstNum + SecondNum
            if (Calc == "-"):
                Resurt = FirstNum - SecondNum
            if (Calc == "×"):
                Resurt = FirstNum * SecondNum
            if (Calc == "÷"):
                Resurt = int((FirstNum / SecondNum) + 0.5)
            await message.channel.send(str(FirstNum) + Calc + str(SecondNum) + "= :thinking:")
            AnswerWait = 1
        if (AnswerWait == 1):
            return message.channel.send(str(Member) + "님이 현재 문제를 풀고 계십니다")

    if message.content.startswith("=인증"):

        Image_captcha = ImageCaptcha()
        msg = ""
        a = ""
        for i in range(6):
            a += str(random.randint(0, 9))

        name = str(message.author.id) + ".png"
        Image_captcha.write(a, name)

        await message.channel.send(file=discord.File(name))
        def check(msg):
            return msg.author == message.author and msg.channel == message.channel

        try:
            msg = await client.wait_for("message", timeout=10, check=check)
        except:
            await message.channel.send("시간 초과")
            return

        if msg.content == a:
            await message.channel.send("인증되었습니다.")
            await message.guild.get_member(message.author.id).add_roles(discord.utils.get(message.guild.roles, name="user"))
        else:
            await message.channel.send("인증에 실패하였습니다. 다시 시도해 주세요.")





        #각각에 print로 로그를 남겨서 작동되는지 확인하고 수정해놓기.
access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
